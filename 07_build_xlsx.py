#!/usr/bin/env python3
"""Build final review XLSX from filtered_with_papers.csv.

Two sheets:
  Candidates — ACCEPT + FLAG rows, sorted ACCEPT-first then by study_title
  Rejected   — REJECT rows (reference / spot-check)

Row color coding:
  Green  = ACCEPT
  Yellow = FLAG
  Red    = REJECT

Review columns (blank — Moshe fills in):
  Soil_or_Rhizo, Rhizosphere_Method, decision (dropdown), notes

Input:  data/filtered_with_papers.csv
Output: candidates_YYYY-MM-DD.xlsx  (in sample_search/ directory)
"""

import os
from datetime import date
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA     = os.path.join(BASE_DIR, "data")
IN_FILE  = os.path.join(DATA, "filtered_with_papers.csv")
OUT_FILE = os.path.join(BASE_DIR, f"candidates_{date.today().isoformat()}.xlsx")

# Columns shown first in the output, in this order
ORDERED_COLS = [
    # ── Identification ────────────────────────────────────────────────────────
    "run_accession", "sample_accession", "secondary_sample_accession",
    "bioproject", "study_accession", "secondary_study_accession",
    "experiment_accession",
    # ── Study / sample content ────────────────────────────────────────────────
    "study_title", "scientific_name", "tax_id",
    "sample_title", "sample_description", "isolation_source", "host",
    # ── Geography ─────────────────────────────────────────────────────────────
    "geo_loc_name", "lat", "lon", "collection_date",
    # ── Sequencing stats ──────────────────────────────────────────────────────
    "library_strategy", "library_source", "library_selection",
    "instrument_platform", "instrument_model",
    "base_count", "read_count",
    "submission_date", "first_public",
    # ── Source & automated QC ────────────────────────────────────────────────
    "_source_db", "auto_quality", "quality_flags",
    # ── Paper context ─────────────────────────────────────────────────────────
    "paper_pmid", "paper_title", "paper_abstract_excerpt",
    # ── Review columns (Moshe fills in) ───────────────────────────────────────
    "Soil_or_Rhizo", "Rhizosphere_Method", "decision", "notes",
]

FILL_ACCEPT = PatternFill("solid", fgColor="C6EFCE")   # green
FILL_FLAG   = PatternFill("solid", fgColor="FFEB9C")   # yellow
FILL_REJECT = PatternFill("solid", fgColor="FFC7CE")   # red
FILL_REVIEW = PatternFill("solid", fgColor="BDD7EE")   # light blue (review cols header)
FILL_HEADER = PatternFill("solid", fgColor="1F4E79")   # dark blue
FONT_HEADER = Font(color="FFFFFF", bold=True, size=10)
FONT_REVIEW = Font(color="1F4E79", bold=True, size=10)

QUALITY_FILL = {"ACCEPT": FILL_ACCEPT, "FLAG": FILL_FLAG, "REJECT": FILL_REJECT}

REVIEW_COLS = {"Soil_or_Rhizo", "Rhizosphere_Method", "decision", "notes"}


def write_sheet(
    wb: openpyxl.Workbook,
    sheet_name: str,
    df: pd.DataFrame,
    is_first: bool = False,
) -> openpyxl.worksheet.worksheet.Worksheet:
    ws = wb.active if is_first else wb.create_sheet(sheet_name)
    ws.title = sheet_name

    cols = list(df.columns)

    # ── Header row ─────────────────────────────────────────────────────────────
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.alignment = Alignment(horizontal="center", wrap_text=False)
        if col in REVIEW_COLS:
            cell.fill = FILL_REVIEW
            cell.font = FONT_REVIEW
        else:
            cell.fill = FILL_HEADER
            cell.font = FONT_HEADER

    # ── Data rows ──────────────────────────────────────────────────────────────
    for ri, (_, row) in enumerate(df.iterrows(), 2):
        q    = str(row.get("auto_quality", ""))
        fill = QUALITY_FILL.get(q)
        for ci, val in enumerate(row, 1):
            cell = ws.cell(
                row=ri, column=ci,
                value=(str(val) if pd.notna(val) and str(val) != "nan" else ""),
            )
            if fill:
                cell.fill = fill

    # ── Freeze top row, set column widths ──────────────────────────────────────
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 20

    for ci, col in enumerate(cols, 1):
        # Sample first 100 rows to estimate needed width
        sample = df.iloc[:100][col].astype(str)
        max_len = max(len(str(col)), sample.str.len().max() if len(sample) else 10)
        # Special widths for known wide columns
        if col in ("paper_abstract_excerpt", "sample_description"):
            width = 60
        elif col in ("study_title", "paper_title", "isolation_source", "quality_flags"):
            width = 45
        elif col in ("Rhizosphere_Method", "notes"):
            width = 35
        elif col in ("run_accession", "sample_accession", "bioproject"):
            width = 18
        else:
            width = min(max_len + 2, 40)
        ws.column_dimensions[get_column_letter(ci)].width = width

    return ws


# ── Main ──────────────────────────────────────────────────────────────────────

df = pd.read_csv(IN_FILE, low_memory=False)
print(f"Loaded {len(df):,} rows")

# Add blank review columns
for col in ["Soil_or_Rhizo", "Rhizosphere_Method", "decision", "notes"]:
    df[col] = ""

# Add any missing ordered columns
for col in ORDERED_COLS:
    if col not in df.columns:
        df[col] = ""

# Reorder: ORDERED_COLS first, then any extras
extra_cols = [c for c in df.columns if c not in ORDERED_COLS]
df = df[ORDERED_COLS + extra_cols]

# Split into Candidates (ACCEPT + FLAG) and Rejected
df_cand = df[df["auto_quality"].isin(["ACCEPT", "FLAG"])].copy()
df_rej  = df[df["auto_quality"] == "REJECT"].copy()

# Sort: ACCEPT first, then FLAG; within each, alphabetical by study_title
sort_key = {"ACCEPT": 0, "FLAG": 1}
df_cand["_sort"] = df_cand["auto_quality"].map(sort_key).fillna(2)
df_cand = df_cand.sort_values(["_sort", "study_title"]).drop(columns=["_sort"])

print(f"Candidates sheet: {len(df_cand):,} rows (ACCEPT + FLAG)")
print(f"Rejected sheet:   {len(df_rej):,} rows")

wb = openpyxl.Workbook()
ws_cand = write_sheet(wb, "Candidates", df_cand, is_first=True)
_        = write_sheet(wb, "Rejected",   df_rej,  is_first=False)

# ── Dropdown validation for 'decision' column ─────────────────────────────────
if "decision" in df_cand.columns:
    dec_idx = list(df_cand.columns).index("decision") + 1
    dec_col = get_column_letter(dec_idx)
    dv = DataValidation(
        type="list",
        formula1='"Good,Rhizosphere,Reject,Review"',
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=False,
    )
    ws_cand.add_data_validation(dv)
    dv.sqref = f"{dec_col}2:{dec_col}{len(df_cand) + 1}"

# Remove the default blank sheet if it snuck in
for name in list(wb.sheetnames):
    if name not in ("Candidates", "Rejected"):
        del wb[name]

wb.save(OUT_FILE)
print(f"\nSaved {OUT_FILE}")
print(f"  Candidates sheet: {len(df_cand):,} rows")
print(f"  Rejected sheet:   {len(df_rej):,} rows")
print()
print("Next steps:")
print("  1. Open the XLSX and work through the Candidates sheet")
print("  2. Fill in 'Soil_or_Rhizo' = Rhizo for confirmed rhizosphere samples")
print("  3. Fill in 'Rhizosphere_Method' from the paper's methods section")
print("  4. Set 'decision' = Good / Reject / Review using the dropdown")
print("  5. Rows marked 'Good' are ready for download and addition to metadata_unified.xlsx")
